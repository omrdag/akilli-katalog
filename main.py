"""
Akıllı Katalog — Bağımsız PDF katalog çıkarma uygulaması.
Ana ByLamp sistemine bağlı DEĞİLDİR. Kendi veritabanı (SQLite) ile çalışır.

Akış:
1. PDF yükle + sayfa + ızgara düzeni seç
2. Sistem ürünleri (SKU, ad, fiyat, görsel) çıkarır
3. Ürünler SQLite'a kaydedilir
4. Akıllı Katalog sayfasında listelenir, aranır, filtrelenir
5. Teklif oluşturulabilir (Excel)
"""
import os
import sqlite3
import io
from datetime import datetime
from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, send_file, g)
from werkzeug.utils import secure_filename

from pdf_extract import extract_catalog_page

app = Flask(__name__)
app.secret_key = os.environ.get('SESSION_SECRET', 'akilli-katalog-dev-key')

UPLOAD_DIR = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_DIR, exist_ok=True)
DB_PATH = os.path.join(os.path.dirname(__file__), 'akilli_katalog.db')


# ── Veritabanı ──
def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DB_PATH)
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop('db', None)
    if db is not None:
        db.close()


def init_db():
    db = sqlite3.connect(DB_PATH)
    db.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sku TEXT,
            product_name TEXT,
            price REAL DEFAULT 0,
            category TEXT,
            image_data TEXT,
            source_pdf TEXT,
            source_page INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    db.commit()
    db.close()


# ── Sayfalar ──
@app.route('/')
def index():
    """Akıllı Katalog ana sayfa — ürünleri listeler, arama/filtre."""
    db = get_db()
    q = (request.args.get('q') or '').strip()
    category = (request.args.get('category') or '').strip()

    sql = 'SELECT * FROM products WHERE 1=1'
    params = []
    if q:
        sql += ' AND (sku LIKE ? OR product_name LIKE ?)'
        params += [f'%{q}%', f'%{q}%']
    if category:
        sql += ' AND category = ?'
        params.append(category)
    sql += ' ORDER BY id DESC'

    products = db.execute(sql, params).fetchall()
    categories = [r['category'] for r in db.execute(
        'SELECT DISTINCT category FROM products WHERE category IS NOT NULL ORDER BY category'
    ).fetchall()]
    total = db.execute('SELECT COUNT(*) c FROM products').fetchone()['c']

    return render_template('index.html', products=products, categories=categories,
                           q=q, category=category, total=total)


@app.route('/upload', methods=['GET', 'POST'])
def upload():
    """PDF yükleme + sayfa/ızgara seçme."""
    if request.method == 'POST':
        file = request.files.get('pdf')
        page = request.form.get('page', type=int)
        category = (request.form.get('category') or '').strip()
        layout_str = (request.form.get('layout') or '3,3,3,4').strip()

        if not file or not file.filename.lower().endswith('.pdf'):
            flash('Lütfen bir PDF dosyası seçin.', 'error')
            return redirect(url_for('upload'))
        if not page or page < 1:
            flash('Geçerli bir sayfa numarası girin.', 'error')
            return redirect(url_for('upload'))

        # Izgara düzenini çöz: "3,3,3,4" -> [3,3,3,4]
        try:
            rows_layout = [int(x) for x in layout_str.split(',') if x.strip()]
            if not rows_layout:
                rows_layout = [3, 3, 3, 4]
        except ValueError:
            rows_layout = [3, 3, 3, 4]

        # Dosyayı kaydet
        filename = secure_filename(file.filename)
        pdf_path = os.path.join(UPLOAD_DIR, filename)
        file.save(pdf_path)

        # Çıkar
        try:
            products = extract_catalog_page(pdf_path, page, rows_layout)
        except Exception as e:
            flash(f'PDF işlenirken hata: {e}', 'error')
            return redirect(url_for('upload'))

        if not products:
            flash('Bu sayfada ürün bulunamadı. Sayfa numarasını kontrol edin.', 'error')
            return redirect(url_for('upload'))

        # Kaydet
        db = get_db()
        saved = 0
        for p in products:
            db.execute('''
                INSERT INTO products (sku, product_name, price, category, image_data, source_pdf, source_page)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                p.get('sku', ''),
                p.get('name', ''),
                float(p.get('price', 0) or 0),
                category or None,
                ('data:image/png;base64,' + p['image_b64']) if p.get('image_b64') else None,
                filename,
                page
            ))
            saved += 1
        db.commit()

        flash(f'{saved} ürün başarıyla çıkarıldı ve kaydedildi.', 'success')
        return redirect(url_for('index'))

    return render_template('upload.html')


@app.route('/delete-all', methods=['POST'])
def delete_all():
    """Tüm çıkarılan ürünleri temizle (test için)."""
    db = get_db()
    db.execute('DELETE FROM products')
    db.commit()
    flash('Tüm ürünler silindi.', 'success')
    return redirect(url_for('index'))


@app.route('/export-quote', methods=['POST'])
def export_quote():
    """Seçili ürünlerden Excel fiyat teklifi oluştur."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    import base64 as b64mod
    import tempfile

    ids = request.form.get('ids', '')
    customer = (request.form.get('customer_name') or '').strip()
    id_list = [int(x) for x in ids.split(',') if x.strip().isdigit()]
    if not id_list:
        flash('Teklif için ürün seçilmedi.', 'error')
        return redirect(url_for('index'))

    db = get_db()
    placeholders = ','.join(['?'] * id_list.__len__())
    rows = db.execute(f'SELECT * FROM products WHERE id IN ({placeholders})', id_list).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Fiyat Teklifi"
    ws.sheet_view.showGridLines = False
    widths = [5, 16, 42, 20, 10, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    DARK = '111111'
    # Başlık
    for col in range(1, 7):
        ws.cell(row=1, column=col).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
        ws.cell(row=2, column=col).fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    ws.merge_cells('A1:D2')
    h = ws.cell(row=1, column=1, value='AKILLI KATALOG — FİYAT TEKLİFİ')
    h.font = Font(size=14, bold=True, color='FFFFFF')
    h.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.merge_cells('E1:F2')
    d = ws.cell(row=1, column=5, value=datetime.now().strftime('%d.%m.%Y'))
    d.font = Font(size=11, color='FFFFFF')
    d.alignment = Alignment(horizontal='right', vertical='center', indent=1)
    ws.row_dimensions[1].height = 22

    if customer:
        ws.cell(row=3, column=1, value=f'Sayın: {customer}').font = Font(size=11, bold=True)

    # Tablo başlığı
    hdr = ['#', 'GÖRSEL', 'ÜRÜN', 'SKU', 'ADET', 'FİYAT']
    for col, t in enumerate(hdr, 1):
        c = ws.cell(row=5, column=col, value=t)
        c.font = Font(size=10, bold=True, color='888888')
        c.border = Border(top=Side(style='medium', color=DARK), bottom=Side(style='medium', color=DARK))

    temp_files = []
    r = 6
    total = 0.0
    for idx, row in enumerate(rows, 1):
        ws.row_dimensions[r].height = 54
        price = float(row['price'] or 0)
        total += price
        ws.cell(row=r, column=1, value=idx)
        # Görsel
        if row['image_data'] and row['image_data'].startswith('data:image'):
            try:
                b64data = row['image_data'].split(',', 1)[1]
                imgbytes = b64mod.b64decode(b64data)
                tf = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                tf.write(imgbytes)
                tf.close()
                temp_files.append(tf.name)
                xi = XLImage(tf.name)
                xi.width = 60
                xi.height = 60
                ws.add_image(xi, f'B{r}')
            except Exception:
                pass
        ws.cell(row=r, column=3, value=row['product_name'] or '')
        ws.cell(row=r, column=4, value=row['sku'] or '')
        ws.cell(row=r, column=5, value=1).alignment = Alignment(horizontal='center')
        pc = ws.cell(row=r, column=6, value=price)
        pc.number_format = '#,##0.00 ₺'
        r += 1

    # Toplam
    r += 1
    tl = ws.cell(row=r, column=5, value='TOPLAM')
    tl.font = Font(bold=True, color='FFFFFF')
    tl.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    tv = ws.cell(row=r, column=6, value=total)
    tv.font = Font(bold=True, color='FFFFFF')
    tv.fill = PatternFill(start_color=DARK, end_color=DARK, fill_type='solid')
    tv.number_format = '#,##0.00 ₺'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    for tf in temp_files:
        try:
            os.unlink(tf)
        except Exception:
            pass

    fname = f"Teklif_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)


init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
